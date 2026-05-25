import openpyxl
from app.extensions import db
from app.models.hospital import Hospital, HospitalCredential
from app.utils.crypto import encrypt


def _clean(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _get_or_create_hospital(name, cache):
    if not name:
        return None
    key = name.lower()
    if key in cache:
        return cache[key]
    hospital = Hospital.query.filter(
        db.func.lower(Hospital.name) == key
    ).first()
    if not hospital:
        hospital = Hospital(name=name, active=True)
        db.session.add(hospital)
        db.session.flush()
    cache[key] = hospital
    return hospital


def _add_cred(hospital, category, label, username=None, password=None,
              host=None, role=None, url=None, notes=None, created_by=None):
    if hospital is None:
        return
    cred = HospitalCredential(
        hospital_id=hospital.id,
        category=category,
        label=label,
        username=username or None,
        password_enc=encrypt(password) if password else None,
        host_enc=encrypt(host) if host else None,
        role_enc=encrypt(role) if role else None,
        url=url or None,
        notes=notes or None,
        created_by=created_by,
    )
    db.session.add(cred)


# ── Sheet handlers ─────────────────────────────────────────────────────────────

def _import_hospital_access(ws, cache, created_by):
    """
    Cols: SITE DESCRIPTION | rustdesk | RustPass | (AnyDesk skip) | (AnydeskPass skip)
          | Os username | OsPass | RemoteDesktop ID | RemoteDesktop Pass
    """
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        site = _clean(row[0])
        if not site:
            continue
        hospital = _get_or_create_hospital(site, cache)

        rustdesk_id = _clean(row[1]) if len(row) > 1 else None
        rust_pass   = _clean(row[2]) if len(row) > 2 else None
        os_user     = _clean(row[5]) if len(row) > 5 else None
        os_pass     = _clean(row[6]) if len(row) > 6 else None
        rd_host     = _clean(row[7]) if len(row) > 7 else None
        rd_pass     = _clean(row[8]) if len(row) > 8 else None

        if rustdesk_id:
            _add_cred(hospital, "remote_desktop", "RustDesk",
                      username=rustdesk_id, password=rust_pass, created_by=created_by)
            count += 1
        if os_user:
            _add_cred(hospital, "os_account", "OS Account",
                      username=os_user, password=os_pass, created_by=created_by)
            count += 1
        if rd_host:
            # rd_host is an IP or hostname — store in host_enc
            _add_cred(hospital, "remote_desktop", "Remote Desktop",
                      host=rd_host, password=rd_pass, created_by=created_by)
            count += 1
    return count


def _import_pacs_ips(ws, cache, created_by):
    """
    Cols: SITE DESCRIPTION | IP ADDRESS | USERNAME | PASSWORD
    Rows with no IP are section headers (hospital name).
    """
    count = 0
    current_hospital = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        site     = _clean(row[0])
        ip       = _clean(row[1]) if len(row) > 1 else None
        username = _clean(row[2]) if len(row) > 2 else None
        password = _clean(row[3]) if len(row) > 3 else None

        if not site:
            continue
        if not ip:
            current_hospital = _get_or_create_hospital(site, cache)
            continue

        hospital = current_hospital or _get_or_create_hospital(site, cache)
        _add_cred(hospital, "network", site,
                  host=ip, username=username, password=password,
                  created_by=created_by)
        count += 1
    return count


def _import_pacs_users(ws, cache, created_by):
    """
    Two-row header — data starts at row 4.
    Col 0: Site | 1-2: Admin Tools (user/pass) | 3-5: PACS Client (user/pass/role)
           | 7-9: WIM (user/pass/role)
    """
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        site = _clean(row[0])
        if not site:
            continue
        hospital = _get_or_create_hospital(site, cache)

        def col(i):
            return _clean(row[i]) if len(row) > i else None

        if col(1):
            _add_cred(hospital, "admin_account", "PACS Admin Tools",
                      username=col(1), password=col(2), created_by=created_by)
            count += 1
        if col(3):
            _add_cred(hospital, "admin_account", "PACS Client",
                      username=col(3), password=col(4), role=col(5),
                      created_by=created_by)
            count += 1
        if col(7):
            _add_cred(hospital, "admin_account", "PACS WIM",
                      username=col(7), password=col(8), role=col(9),
                      created_by=created_by)
            count += 1
    return count


def _import_pyxis_ips(ws, cache, created_by):
    """
    Cols: SITE DESCRIPTION | IP | USERNAME | PASSWORD | (blank) | Application URL | USERNAME | PASSWORD
    """
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        site = _clean(row[0])
        if not site:
            continue
        hospital = _get_or_create_hospital(site, cache)

        def col(i):
            return _clean(row[i]) if len(row) > i else None

        if col(1):
            _add_cred(hospital, "network", "Pyxis IP",
                      host=col(1), username=col(2), password=col(3),
                      created_by=created_by)
            count += 1
        if col(5) or col(6):
            _add_cred(hospital, "admin_account", "Pyxis Application",
                      url=col(5), username=col(6), password=col(7),
                      created_by=created_by)
            count += 1
    return count


def _import_pyxis_users(ws, cache, created_by):
    """
    Two-row header — data starts at row 3.
    Col 0: Site | 1: Username | 2: Password | 3: Role | 4: Notes/server
    Site is blank on continuation rows — carry forward.
    """
    count = 0
    current_site = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        site     = _clean(row[0])
        username = _clean(row[1]) if len(row) > 1 else None
        password = _clean(row[2]) if len(row) > 2 else None
        role     = _clean(row[3]) if len(row) > 3 else None
        notes    = _clean(row[4]) if len(row) > 4 else None

        if site:
            current_site = site
        if not username or not current_site:
            continue

        hospital = _get_or_create_hospital(current_site, cache)
        _add_cred(hospital, "admin_account", "Pyxis User",
                  username=username, password=password, role=role,
                  notes=notes, created_by=created_by)
        count += 1
    return count


def _import_cvis_ips(ws, cache, created_by):
    """
    Cols: Server Name | IP ADDRESS | USERNAME | PASSWORD
    Rows with no IP are section headers (hospital name).
    """
    count = 0
    current_hospital = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        server   = _clean(row[0])
        ip       = _clean(row[1]) if len(row) > 1 else None
        username = _clean(row[2]) if len(row) > 2 else None
        password = _clean(row[3]) if len(row) > 3 else None

        if not server:
            continue
        if not ip:
            current_hospital = _get_or_create_hospital(server, cache)
            continue

        hospital = current_hospital or _get_or_create_hospital(server, cache)
        _add_cred(hospital, "network", f"CVIS - {server}",
                  host=ip, username=username, password=password,
                  created_by=created_by)
        count += 1
    return count


def _import_vpn(ws, cache, created_by):
    """
    Cols: VPN type | Connection name | Remote Gateway | Port | USERNAME | Password
    VPN type and connection name carry forward on blank rows.
    """
    count = 0
    current_vpn_type = None
    current_connection = None
    current_gateway = None
    current_port = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        def col(i):
            return _clean(row[i]) if len(row) > i else None

        if col(0):
            current_vpn_type = col(0)
        if col(1):
            current_connection = col(1)
        if col(2):
            current_gateway = col(2)
        if col(3) is not None and str(row[3]).strip():
            current_port = _clean(row[3])

        username = col(4)
        password = col(5)

        if not username or not current_connection:
            continue

        hospital = _get_or_create_hospital(current_connection, cache)

        note_parts = []
        if current_vpn_type:
            note_parts.append(current_vpn_type)
        if current_gateway:
            port_str = f":{current_port}" if current_port else ""
            note_parts.append(f"{current_gateway}{port_str}")

        _add_cred(hospital, "vpn", f"VPN - {current_connection}",
                  username=username, password=password,
                  notes=" | ".join(note_parts) if note_parts else None,
                  created_by=created_by)
        count += 1
    return count


# ── Main entry point ───────────────────────────────────────────────────────────

_SHEET_HANDLERS = {
    "hospital access": _import_hospital_access,
    "pacs ips":        _import_pacs_ips,
    "pacs users":      _import_pacs_users,
    "pyxis ips":       _import_pyxis_ips,
    "pyxis users":     _import_pyxis_users,
    "cvis ips":        _import_cvis_ips,
    "vpn":             _import_vpn,
}


def import_sites_excel(filepath, created_by=None):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Build a normalized lookup so leading/trailing spaces in sheet names don't break matching
    sheet_lookup = {name.strip().lower(): name for name in wb.sheetnames}

    hospital_cache = {}
    sheet_stats = {}
    total_creds = 0

    for key, handler in _SHEET_HANDLERS.items():
        real_name = sheet_lookup.get(key)
        if not real_name:
            sheet_stats[key] = {"skipped": True}
            continue
        ws = wb[real_name]
        count = handler(ws, hospital_cache, created_by)
        sheet_stats[key] = {"credentials": count}
        total_creds += count

    db.session.commit()

    return {
        "hospitals": len(hospital_cache),
        "credentials": total_creds,
        "sheets": sheet_stats,
    }
